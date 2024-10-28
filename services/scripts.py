import openpyxl
from .cfg import RegIndex,TempCell,Abbreviations,Abbreviations_Decryption,RegIndexGov
import os.path
from time import sleep
from sys import exit
from progress.bar import Bar
import random

def startup_check():
        print('Запуск программы...')
        sleep(0.25)
        print('Проверка файлов...\n')
        sleep(0.25)
        if os.path.exists('./import_files/template.xlsx'):
            pass
        else:
            print('Ошибка! Не обнаружен файл шаблона. Проверьте наличие файла "template.xlsx" в папке "import_files". \n Программа будет завершена через 10 секунд.')
            sleep(12)
            exit()
        if not os.path.isdir('./print_files'):
            os.mkdir('./print_files')
        if not os.path.isdir('./export_files'):
            os.mkdir('./export_files')
        if not os.path.isdir('./randomize'):
            os.mkdir('./randomize')

def registry_input():
        while True:
            registry_filename = str(input('Введите название файла реестра (Пример: somename): '))
            if os.path.exists(f'./import_files/{registry_filename}.xlsx'):
                break
            else:
                print('\n Ошибка! Данный файл не найден. \n Проверьте название файла, уберите лишние пробелы и/или перезапустите программу. \n')
        print('')
        return registry_filename

# Считывание реестра-excel в список [[компания1],[компания2]]
def registry_read(filename: str) -> list:
    book = openpyxl.open(f'./import_files/{filename}.xlsx',read_only=True)
    registry = book.active
    reg_list = []
    for row in registry.iter_rows(min_row=2,max_row=registry.max_row,min_col=registry.min_column,max_col=registry.max_column+2):
        company = []
        for cell in row:
            if cell.value != None:
                company.append(cell.value)
            else:
                company.append('Нет Данных')
        reg_list.append(company)
    book.close()
    print(f'<><><><><>\nРеестр успешно прочтен \n Количество компаний: {len(reg_list)}\n<><><><><>\n')
    sleep(0.2)
    return reg_list

# Создает обложку для компании сответсвенно шаблону ЧАСТНИК
def create_companies_sheet_private(companies: list, template_filename: str) -> None:
    print('<><><><><>\nНачало обработки файлов\n<><><><><>\n')
    working_time = Bar('Обработка файлов',max=len(companies),suffix='%(percent).1f%% - %(eta)ds')
    book = openpyxl.open(f'./import_files/{template_filename}.xlsx')
    company_head_sheet = book.active

    for company in companies:
        company_name = (company[RegIndex.NAME].split())
        if company_name[0] in Abbreviations:
            company_head_sheet[TempCell.NAME] = Abbreviations_Decryption[(Abbreviations.index(company_name[0]))]+' '+''.join(company_name[1:])
        else:
            company_head_sheet[TempCell.NAME] = company[RegIndex.NAME]
        company_head_sheet[TempCell.LEGAL_ADDRESS] = company[RegIndex.ADDRESS]
        company_head_sheet[TempCell.POSTMAIL_ADDRESS] = company[RegIndex.ADDRESS]
        company_head_sheet[TempCell.ADMINISTRATOR] = f'{company[RegIndex.SURNAME]} {company[RegIndex.FIRSTNAME]} {company[RegIndex.PATRONYMIC]}'
        company_head_sheet[TempCell.YEAR] = '2024'
        company_head_sheet[TempCell.OWNERSHIP_FORM] = 'Частная собственность' 
        company_head_sheet[TempCell.OKPO] = company[RegIndex.OKPO]
        company_head_sheet[TempCell.OKVED_ACTIVITY_TYPE] = company[RegIndex.OKVED_ACTIVITY_TYPE]
        company_head_sheet[TempCell.OKATO_TERRITORY] = company[RegIndex.OKATO_TERRITORY]
        company_head_sheet[TempCell.INN] = company[RegIndex.INN]
        company_head_sheet[TempCell.KPP] = company[RegIndex.KPP]
        company_head_sheet[TempCell.OGRN] = company[RegIndex.OGRN]
        fixed_name_for_save = str(company[RegIndex.INN])
        book.save('./export_files/'+f'{fixed_name_for_save}.xlsx')
        working_time.next()
    working_time.finish()


    book.close()
    print('\n<><><><><>\nВсе файлы успешно обработаны\nПроверьте папку export_files\n<><><><><>\n')

# Создает обложку для компании сответсвенно шаблону ГОС
def create_companies_sheet_gov(companies: list, template_filename: str) -> None:
    print('<><><><><>\nНачало обработки файлов\n<><><><><>\n')
    working_time = Bar('Обработка файлов',max=len(companies),suffix='%(percent).1f%% - %(eta)ds')
    book = openpyxl.open(f'./import_files/{template_filename}.xlsx')
    company_head_sheet = book.active
    for company in companies:
        company_name = (company[RegIndex.NAME].split())
        if company_name[0] in Abbreviations:
            company_head_sheet[TempCell.NAME] = Abbreviations_Decryption[(Abbreviations.index(company_name[0]))]+' '+''.join(company_name[1:])
        else:
            company_head_sheet[TempCell.NAME] = company[RegIndexGov.NAME]
        company_head_sheet[TempCell.LEGAL_ADDRESS] = company[RegIndexGov.ADDRESS]
        company_head_sheet[TempCell.POSTMAIL_ADDRESS] = company[RegIndexGov.ADDRESS]
        company_head_sheet[TempCell.ADMINISTRATOR] = f'{company[RegIndexGov.SURNAME]} {company[RegIndexGov.FIRSTNAME]} {company[RegIndexGov.PATRONYMIC]}'
        company_head_sheet[TempCell.YEAR] = '2024'
        company_head_sheet[TempCell.OWNERSHIP_FORM] = 'Государственная собственность' 
        company_head_sheet[TempCell.OKPO] = company[RegIndexGov.OKPO]
        company_head_sheet[TempCell.OKVED_ACTIVITY_TYPE] = company[RegIndexGov.OKVED_ACTIVITY_TYPE]
        company_head_sheet[TempCell.OKATO_TERRITORY] = company[RegIndexGov.OKATO_TERRITORY]
        company_head_sheet[TempCell.INN] = company[RegIndexGov.INN]
        company_head_sheet[TempCell.KPP] = company[RegIndexGov.KPP]
        company_head_sheet[TempCell.OGRN] = company[RegIndexGov.OGRN]
        fixed_name_for_save = str(company[RegIndexGov.INN])
        book.save('./export_files/'+f'{fixed_name_for_save}.xlsx')
        working_time.next()
    working_time.finish()


    book.close()
    print('\n<><><><><>\nВсе файлы успешно обработаны\nПроверьте папку export_files\n<><><><><>\n')

# Переводит таблицу excel в нормализованный вид
def printing_fix():
    if not os.path.isdir('./export_files/print'):
        os.mkdir('./export_files/print')
    print('<><><><><>\nНачало обработки файлов\n<><><><><>\n')
    amount = 0
    for filename in os.listdir('./print_files'):
        amount += 1
    working_time = Bar('Обработка файлов',max=amount,suffix='%(percent).1f%% - %(eta)ds')
    for filename in os.listdir('./print_files'):
        if os.path.splitext(filename)[1] in ['.xlsx','.xlsm','.xltx','.xltm']:
            book = openpyxl.open(f'./print_files/{filename}')
            sheet = book.active
            sheet.column_dimensions['A'].width = 32
            sheet.column_dimensions['B'].width = 15
            for i in ['C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB']:
                sheet.column_dimensions[i].width = 7
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToHeight = False
            sheet.page_setup.fitToWidth = 1
            book.save('./export_files/print/'+f'{filename}')
            book.close()
            
        else:
            if not os.path.isdir('./export_files/print_errors'):
                os.mkdir('./export_files/print_errors')
            os.rename(f'./print_files/{filename}',f'./export_files/print_errors/{filename}')
        working_time.next()
    working_time.finish()
    print('\n<><><><><>\nВсе файлы успешно обработаны\nПроверьте папку export_files -> print\nВсе файлы с не поддерживаемым типом файла находятся в export_files -> print_errors\n<><><><><>\n')

# рандомайзит любые файлы в другие названия
def random_naming():
    amount = 0
    for _ in os.listdir('./randomize'):
        amount += 1
    working_time = Bar('Рандомизация названий',max=amount,suffix='%(percent).1f%% - %(eta)ds')
    alredy_present = []
    if amount == 0:
        print('Ошибка! Не обнаружены файлы для рандомизации названий в папке "randomize"!')
    for filename in os.listdir('./randomize'):
        while True:
            random_name = random.randint(10**9,10**18)
            if random_name not in alredy_present:
                alredy_present.append(random_name)
                os.rename(f'./randomize/{filename}',f'./randomize/{random_name}{os.path.splitext(f'./randomize/{filename}')[1]}')
                break
        working_time.next()
    working_time.finish()
    

